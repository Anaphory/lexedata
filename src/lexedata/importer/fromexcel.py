# -*- coding: utf-8 -*-

import re
import argparse
import typing as t
from pathlib import Path
from tempfile import mkdtemp
import logging
from collections import OrderedDict

from tqdm import tqdm

import pycldf
import openpyxl

from lexedata.types import (
    Object,
    Language,
    RowObject,
    Form,
    Concept,
    CogSet,
    Judgement,
)
from lexedata.util import (
    string_to_id,
    clean_cell_value,
    get_cell_comment,
    edit_distance,
)
import lexedata.importer.cellparser as cell_parsers
import lexedata.error_handling as err


Ob = t.TypeVar("O", bound=Object)

logger = logging.getLogger(__name__)

# NOTE: Excel uses 1-based indices, this shows up in a few places in this file.


def cells_are_empty(cells: t.Iterable[openpyxl.cell.Cell]) -> bool:
    return not any([clean_cell_value(cell) for cell in cells])


class DB:
    cache: t.Dict[str, t.Dict[t.Hashable, t.Dict[str, t.Any]]]

    def __init__(self, output_dataset: pycldf.Wordlist, fname=None):
        if fname is not None:
            logger.info("Warning: dbase fname set, but ignored")
        self.dataset = output_dataset
        self.cache = {}
        self.source_ids = set()

    def cache_dataset(self):
        for table in self.dataset.tables:
            table_type = (
                table.common_props.get("dc:conformsTo", "").rsplit("#", 1)[1]
                or table.url
            )
            (id,) = table.tableSchema.primaryKey
            self.cache[table_type] = {row[id]: row for row in table}
        # TODO: what is the actual API
        for source in self.dataset.sources:
            self.source_ids.add(source.id)

    def drop_from_cache(self, table: str):
        self.cache[table] = {}

    def retrieve(self, table_type: str):
        return self.cache[table_type].values()

    def add_source(self, source_id):
        self.source_ids.add(source_id)

    def empty_cache(self):
        self.cache = {
            # TODO: Is there a simpler way to get the list of all tables?
            table.common_props.get("dc:conformsTo", "").rsplit("#", 1)[1]
            or table.url: {}
            for table in self.dataset.tables
        }

    def write_dataset_from_cache(self, tables: t.Optional[t.List[str]] = None):
        if tables is None:
            tables = self.cache.keys()
        for table_type in tables:
            self.dataset[table_type].common_props["dc:extent"] = self.dataset[
                table_type
            ].write(self.retrieve(table_type))
        self.dataset.write_metadata()
        # TODO: Write BIB file, without pycldf
        for source in self.source_ids:
            print("@misc{" + source + ", title={" + source + "} }")

    def associate(
        self, form_id: str, row: RowObject, comment: t.Optional[str] = None
    ) -> bool:
        form = self.cache["FormTable"][form_id]
        if row.__table__ == "CognatesetTable":
            id = self.dataset["CognatesetTable", "id"].name
            try:
                column = self.dataset["FormTable", "cognatesetReference"]
            except KeyError:
                judgements = self.cache["CognateTable"]
                cognateset = row[self.dataset["CognatesetTable", "id"].name]
                judgement = Judgement(
                    {
                        self.dataset["CognateTable", "id"].name: "{:}-{:}".format(
                            form_id, cognateset
                        ),
                        self.dataset["CognateTable", "formReference"].name: form_id,
                        self.dataset[
                            "CognateTable", "cognatesetReference"
                        ].name: cognateset,
                        self.dataset["CognateTable", "comment"].name: comment or "",
                    }
                )
                self.make_id_unique(judgement)
                judgements[judgement[id]] = judgement
                return True
        elif row.__table__ == "ParameterTable":
            column = self.dataset["FormTable", "parameterReference"]
            id = self.dataset["ParameterTable", "id"].name

        if column.separator is None:
            form[column.name] = row[id]
        else:
            form.setdefault(column.name, []).append(row[id])
        return True

    def insert_into_db(self, object: Ob) -> bool:
        id = self.dataset[object.__table__, "id"].name
        assert object[id] not in self.cache[object.__table__]
        self.cache[object.__table__][object[id]] = object
        return True

    def make_id_unique(self, object: Ob) -> str:
        id = self.dataset[object.__table__, "id"].name
        raw_id = object[id]
        i = 0
        while object[id] in self.cache[object.__table__]:
            i += 1
            object[id] = "{:}_{:d}".format(raw_id, i)
        return object[id]

    def find_db_candidates(
        self,
        object: Ob,
        properties_for_match: t.Iterable[str],
        edit_dist_threshold: t.Optional[int] = None,
    ) -> t.Iterable[str]:
        if edit_dist_threshold:

            def match(x, y):
                return edit_distance(x, y) <= edit_dist_threshold

        else:

            def match(x, y):
                return x == y
        #candidates = [candidate for candidate, properties in self.cache[object.__table__].items() if
                      # all(match(properties[p], object[p]) for p in properties)]
        candidates = []
        for candidate, properties in self.cache[object.__table__].items():
            properties_in_object = [p for p in properties_for_match if p in object and p in properties]
            if all(match(properties[p], object[p]) for p in properties_in_object):
                candidates.append(candidate)
        return candidates


    def commit(self):
        pass


class ExcelParser:
    def __init__(
        self,
        output_dataset: pycldf.Dataset,
        db_fname: str,
        row_object: Object = Concept,
        top: int = 2,
        cellparser: cell_parsers.NaiveCellParser = cell_parsers.CellParser,
        # The following column names should be generated from CLDF terms. This
        # will likely mean that the __init__ method would have to get a
        # slightly different signature, to take that dependency on the output
        # dataset into account.
        row_header: t.List[str] = ["set", "Name", None],
        check_for_match: t.List[str] = ["ID"],
        check_for_row_match: t.List[str] = ["Name"],
        check_for_language_match: t.List[str] = ["Name"],
        on_language_not_found: err.MissingHandler = err.create,
        on_row_not_found: err.MissingHandler = err.create,
        on_form_not_found: err.MissingHandler = err.create,
    ) -> None:
        self.on_language_not_found = on_language_not_found
        self.on_row_not_found = on_row_not_found
        self.on_form_not_found = on_form_not_found
        self.row_header = row_header
        try:
            self.cell_parser = getattr(cell_parsers, cellparser)(output_dataset)
        except TypeError:
            self.cell_parser = cellparser
        self.row_object = row_object
        self.top = top
        self.left = len(row_header) + 1
        self.check_for_match = check_for_match
        self.check_for_row_match = check_for_row_match
        self.check_for_language_match = check_for_language_match
        self.db = DB(output_dataset, fname=db_fname)

    def language_from_column(self, column: t.List[openpyxl.cell.Cell]) -> Language:
        data = [clean_cell_value(cell) for cell in column[: self.top - 1]]
        comment = get_cell_comment(column[0])
        id = string_to_id(data[0])
        return Language(
            # an id candidate must be provided, which is transformed into a unique id
            ID=id,
            Name=data[0],
            Comment=comment,
        )

    def properties_from_row(
        self, row: t.List[openpyxl.cell.Cell]
    ) -> t.Optional[RowObject]:
        c_id = self.db.dataset[self.row_object.__table__, "id"].name
        c_comment = self.db.dataset[self.row_object.__table__, "comment"].name
        c_name = self.db.dataset[self.row_object.__table__, "name"].name
        data = [clean_cell_value(cell) for cell in row[: self.left - 1]]
        properties = dict(zip(self.row_header, data))
        # delete all possible None entries coming from row_header
        while None in properties.keys():
            del properties[None]

        # fetch cell comment
        comment = get_cell_comment(row[0])
        properties[c_comment] = comment

        # cldf_name serves as cldf_id candidate
        properties[c_id] = properties[c_name]

        return self.row_object(properties)

    def parse_all_languages(
        self, sheet: openpyxl.worksheet.worksheet.Worksheet
    ) -> t.Dict[str, str]:
        """Parse all language descriptions in the focal sheet.

        Returns
        =======
        languages: A dictionary mapping columns ("B", "C", "D", …) to language IDs
        """
        languages_by_column: t.Dict[str, str] = {}
        # iterate over language columns
        for lan_col in tqdm(
            sheet.iter_cols(min_row=1, max_row=self.top - 1, min_col=self.left),
            total=sheet.max_column - self.left,
        ):
            if cells_are_empty(lan_col):
                # Skip empty languages
                continue
            language = self.language_from_column(lan_col)
            candidates = self.db.find_db_candidates(
                language, self.check_for_language_match
            )
            for language_id in candidates:
                break
            else:
                if not (
                    self.on_language_not_found(language, lan_col[0])
                    and self.db.insert_into_db(language)
                ):
                    continue
                # TODO: use term→column name
                language_id = language["ID"]
            languages_by_column[lan_col[0].column] = language_id

        return languages_by_column

    def parse_cells(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        languages = self.parse_all_languages(sheet)
        row_object = None
        c_f_id = self.db.dataset["FormTable", "id"].name
        c_f_language = self.db.dataset["FormTable", "languageReference"].name
        c_f_value = self.db.dataset["FormTable", "value"].name
        c_f_comment = self.db.dataset["FormTable", "comment"].name
        c_f_source = self.db.dataset["FormTable", "source"].name
        for row in tqdm(
            sheet.iter_rows(min_row=self.top), total=sheet.max_row - self.top
        ):
            row_header, row_forms = row[: self.left - 1], row[self.left - 1 :]
            # Parse the row header, creating or retrieving the associated row
            # object (i.e. a concept or a cognateset)
            properties = self.properties_from_row(row_header)
            if properties:
                c_r_id = self.db.dataset[properties.__table__, "id"].name
                c_r_name = self.db.dataset[properties.__table__, "name"].name
                similar = self.db.find_db_candidates(
                    properties, self.check_for_row_match
                )
                for row_id in similar:
                    properties[c_r_id] = row_id
                    break
                else:
                    if self.on_row_not_found(properties, row[0]):
                        if c_r_id not in properties:
                            properties[c_r_id] = string_to_id(
                                str(properties.get(c_r_name, ""))
                            )
                        self.db.make_id_unique(properties)
                        self.db.insert_into_db(properties)
                    else:
                        continue
                row_object = properties

            if row_object is None:
                if any(c.value for c in row_forms):
                    raise AssertionError(
                        "Empty first row: Row had no properties, "
                        "and there was no previous row to copy"
                    )
                else:
                    continue
            # Parse the row, cell by cell
            for cell_with_forms in row_forms:
                try:
                    this_lan = languages[cell_with_forms.column]
                except KeyError:
                    continue

                # Parse the cell, which results (potentially) in multiple forms
                for params in self.cell_parser.parse(
                    cell_with_forms,
                    this_lan,
                    f"{sheet.title}.{cell_with_forms.coordinate}",
                ):
                    if c_f_source in params:
                        # Add all sources to sources set
                        for source, _ in params[c_f_source]:
                            self.db.add_source(source)
                        # Glue contex to source
                        params[c_f_source] = {
                            f"{source:}[{context:}]" if context else source
                            for source, context in params[c_f_source]
                        }
                    # Cellparser adds comment of a excel cell to "Cell_Comment" if given
                    # TODO: Gereon, I don't think that params.pop(c_f_comment, None) is correct here. Is it?
                    maybe_comment: t.Optional[str] = params.pop(c_f_comment, None)
                    form = Form(params)
                    if c_f_id not in form:
                        # create candidate for form[id]
                        form[c_f_id] = "{:}_{:}".format(
                            form[c_f_language], row_object[c_r_id]
                        )
                    candidate_forms = iter(self.db.find_db_candidates(form, self.check_for_match))
                    try:
                        # if a candidate for form already exists, don't add the form
                        form_id = next(candidate_forms)
                        self.db.associate(form_id, row_object, maybe_comment)
                    except StopIteration:
                        # no candidates. form is created or not.
                        if self.on_form_not_found(form, cell_with_forms):
                            form[c_f_id] = "{:}_{:}".format(
                                form[c_f_language], row_object[c_r_id]
                            )
                            self.db.insert_into_db(form)
                            form_id = form[c_f_id]
                            self.db.associate(
                                form_id, row_object, comment=maybe_comment
                            )
                        else:
                            logger.error(
                                "The missing form was {:} in {:}, given as {:}.".format(
                                    row_object[c_r_id], this_lan, form[c_f_value]
                                )
                            )
                            # TODO: Fill data with a fuzzy search
                            for row in self.db.find_db_candidates(
                                form, self.check_for_match, edit_dist_threshold=4
                            ):
                                logger.info(f"Did you mean {row} ?")
                            continue
        self.db.commit()


class ExcelCognateParser(ExcelParser):
    def __init__(
        self,
        output_dataset: pycldf.Dataset,
        db_fname: str,
        row_object: Object = CogSet,
        top: int = 2,
        cellparser: cell_parsers.NaiveCellParser = cell_parsers.CellParser,
        row_header=["set", "Name", None],
        check_for_match: t.List[str] = ["ID"],
        check_for_row_match: t.List[str] = ["Name"],
        check_for_language_match: t.List[str] = ["Name"],
        on_language_not_found: err.MissingHandler = err.error,
        on_row_not_found: err.MissingHandler = err.create,
        on_form_not_found: err.MissingHandler = err.warn,
    ) -> None:
        super().__init__(
            output_dataset=output_dataset,
            db_fname=db_fname,
            row_object=row_object,
            top=top,
            cellparser=cellparser,
            check_for_match=check_for_match,
            row_header=row_header,
            check_for_row_match=check_for_row_match,
            check_for_language_match=check_for_language_match,
            on_language_not_found=on_language_not_found,
            on_row_not_found=on_row_not_found,
            on_form_not_found=on_form_not_found,
        )

    def properties_from_row(
        self, row: t.List[openpyxl.cell.Cell]
    ) -> t.Optional[RowObject]:
        # TODO: Ask Gereon: get_cell_comment with unicode normalization or not?
        c_id = self.db.dataset[self.row_object.__table__, "id"].name
        c_comment = self.db.dataset[self.row_object.__table__, "comment"].name
        c_name = self.db.dataset[self.row_object.__table__, "name"].name
        data = [clean_cell_value(cell) for cell in row[: self.left - 1]]
        properties = dict(zip(self.row_header, data))
        # delete all possible None entries coming from row_header
        while None in properties.keys():
            del properties[None]

        # fetch cell comment
        comment = get_cell_comment(row[0])
        properties[c_comment] = comment

        # cldf_name serves as cldf_id candidate
        properties[c_id] = properties.get(c_id) or properties[c_name]
        return self.row_object(properties)

    def associate(
        self, form_id: str, row: RowObject, comment: t.Optional[str] = None
    ) -> bool:
        c_id = self.db.dataset[self.row_object, "id"].name
        assert (
            row.__table__ == "CognatesetTable"
        ), "Expected CognateSet, but got {:}".format(row.__class__)
        row_id = row[c_id]
        judgement = Judgement(
            cldf_id=f"{form_id}-{row_id}",
            cldf_formReference=form_id,
            cldf_cognatesetReference=row_id,
            cldf_comment=comment or "",
        )
        self.db.make_id_unique(judgement)
        return self.db.insert_into_db(judgement)


def excel_parser_from_dialect(output_dataset,
    dialect: argparse.Namespace, cognate: bool
) -> t.Type[ExcelParser]:
    if cognate:
        Row: t.Type[RowObject] = CogSet
        Parser: t.Type[ExcelParser] = ExcelCognateParser
    else:
        Row = Concept
        Parser = ExcelParser
    top = len(dialect.lang_cell_regexes) + 1
    # prepare cellparser
    row_header = []
    for row_regex in dialect.row_cell_regexes:
        match = re.fullmatch(row_regex, "", re.DOTALL)
        row_header += list(match.groupdict().keys()) or [None]
    element_semantics = OrderedDict()
    bracket_pairs = OrderedDict()
    for value in dialect.cell_parser["cell_parser_semantics"]:
        name, opening, closing, my_bool = value
        bracket_pairs[opening] = closing
        element_semantics[opening] = (name, my_bool)
    initialized_cell_parser = getattr(cell_parsers, dialect.cell_parser["name"])(
        output_dataset,
        bracket_pairs=bracket_pairs,
        element_semantics=element_semantics,
        separation_pattern=fr"([{''.join(dialect.cell_parser['form_separator'])}])",
        variant_separator=dialect.cell_parser["variant_separator"],
        add_default_source=dialect.cell_parser.get("add_default_source"),
    )

    class SpecializedExcelParser(Parser):
        def __init__(
            self,
            output_dataset: pycldf.Dataset,
            db_fname: str,
        ) -> None:
            super().__init__(
                output_dataset=output_dataset,
                db_fname=db_fname,
                top=top,
                row_object=Row,
                row_header=row_header,
                cellparser=initialized_cell_parser,
                check_for_match=dialect.check_for_match,
                check_for_row_match=dialect.check_for_row_match,
                check_for_language_match=dialect.check_for_language_match,
            )
            self.db.empty_cache()

        def language_from_column(self, column: t.List[openpyxl.cell.Cell]) -> Language:
            """Parse the row, according to regexes from the metadata.

            Raises
            ======
            ValueError: When the cell cannot be parsed with the specified regex.

            """
            d: t.Dict[str, str] = {}
            for cell, cell_regex, comment_regex in zip(
                column, dialect.lang_cell_regexes, dialect.lang_comment_regexes
            ):
                if cell.value:
                    match = re.fullmatch(cell_regex, cell.value.strip(), re.DOTALL)
                    if match is None:
                        raise ValueError(
                            f"In cell {cell.coordinate}: Expected to encounter match "
                            f"for {cell_regex}, but found {cell.value}"
                        )
                    for k, v in match.groupdict().items():
                        if k in d:
                            d[k] = d[k] + v
                        else:
                            d[k] = v
                if cell.comment:
                    match = re.fullmatch(comment_regex, cell.comment.content, re.DOTALL)
                    if match is None:
                        raise ValueError(
                            f"In cell {cell.coordinate}: Expected to encounter match "
                            f"for {comment_regex}, but found {cell.comment.content}"
                        )
                    for k, v in match.groupdict().items():
                        if k in d:
                            d[k] = d[k] + v
                        else:
                            d[k] = v

            c_l_id = self.db.dataset["LanguageTable", "id"].name
            c_l_name = self.db.dataset["LanguageTable", "name"].name
            if c_l_id not in d:
                d[c_l_id] = string_to_id(d[c_l_name])
            return Language(d)

        def properties_from_row(self, row: t.List[openpyxl.cell.Cell]) -> Row:
            """Parse the row, according to regexes from the metadata.

            Raises
            ======
            ValueError: When the cell cannot be parsed with the specified regex.

            """
            d: t.Dict[str, str] = {}
            for cell, cell_regex, comment_regex in zip(
                row, dialect.row_cell_regexes, dialect.row_comment_regexes
            ):
                if cell.value:
                    match = re.fullmatch(cell_regex, cell.value.strip(), re.DOTALL)
                    if match is None:
                        raise ValueError(
                            f"In cell {cell.coordinate}: Expected to encounter match"
                            f"for {cell_regex}, but found {cell.value}"
                        )
                    for k, v in match.groupdict().items():
                        if k in d:
                            d[k] = d[k] + v
                        else:
                            d[k] = v
                if cell.comment:
                    match = re.fullmatch(comment_regex, cell.comment.content, re.DOTALL)
                    if match is None:
                        raise ValueError(
                            f"In cell {cell.coordinate}: Expected to encounter match "
                            f"for {comment_regex}, but found {cell.comment.content}"
                        )
                    for k, v in match.groupdict().items():
                        if k in d:
                            d[k] = d[k] + v
                        else:
                            d[k] = v

            return Row(d)

    return SpecializedExcelParser


def load_dataset(
    metadata: Path, lexicon: str, db: Path, cognate_lexicon: t.Optional[str]
):
    # logging.basicConfig(filename="warnings.log")
    dataset = pycldf.Dataset.from_metadata(metadata)
    if db == "":
        tmpdir = Path(mkdtemp("", "fromexcel"))
        db = tmpdir / "db.sqlite"
    lexicon_wb = openpyxl.load_workbook(lexicon).active
    # load dialect from metadata
    try:
        dialect = argparse.Namespace(**dataset.tablegroup.common_props["special:fromexcel"])
    except KeyError:
        logger.warning(
            "Dialect not found or dialect was missing a key, "
            "falling back to default parser"
        )
        dialect = None
    # load lexical data set
    if dialect:
        EP = excel_parser_from_dialect(dataset, dialect, cognate=False)
    else:
        EP = ExcelParser
    # The Intermediate Storage, in a in-memory DB (unless specified otherwise)
    EP = EP(dataset, db_fname=db)
    EP.db.empty_cache()
    EP.parse_cells(lexicon_wb)
    # load cognate data set if provided by metadata
    cognate = True if cognate_lexicon and dialect.cognates else False
    if cognate:
        ECP = excel_parser_from_dialect(
            dataset, argparse.Namespace(**dialect.cognates), cognate=cognate
        )
    else:
        ECP = ExcelCognateParser
    ECP = ECP(dataset, db)
    ECP.db = EP.db
    if cognate_lexicon:  # in case no cognate file given, nothing happens
        for sheet in openpyxl.load_workbook(cognate_lexicon).worksheets:
            ECP.parse_cells(sheet)
    ECP.db.write_dataset_from_cache()


if __name__ == "__main__":
    import argparse
    import pycldf

    parser = argparse.ArgumentParser(
        description="Load a Maweti-Guarani-style dataset into CLDF"
    )
    parser.add_argument(
        "lexicon",
        nargs="?",
        default="TG_comparative_lexical_online_MASTER.xlsx",
        help="Path to an Excel file containing the dataset",
    )
    parser.add_argument(
        "--cogsets",
        nargs="?",
        default="",
        help="Path to an Excel file containing cogsets and cognatejudgements",
    )
    parser.add_argument(
        "--db",
        nargs="?",
        default="",
        help="Where to store the temp from reading the word list",
    )
    parser.add_argument(
        "--metadata",
        nargs="?",
        type=Path,
        default="Wordlist-metadata.json",
        help="Path to the metadata.json",
    )
    parser.add_argument(
        "--debug-level",
        type=int,
        default=0,
        help="Debug level: Higher numbers are less forgiving",
    )
    args = parser.parse_args()

    if args.db.startswith("sqlite:///"):
        args.db = args.db[len("sqlite:///") :]
    if args.db == ":memory:":
        args.db = ""

    load_dataset(args.metadata, args.lexicon, args.db, args.cogsets)
